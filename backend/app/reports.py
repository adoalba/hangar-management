"""
Aviation Inventory Granular Reports Module
FAA/EASA Compliance Style Reporting API

This module provides endpoints for generating detailed inventory reports
with cross-filtering capabilities and audit-grade metadata.
"""

from flask import Blueprint, request, jsonify, g, send_file, Response
from sqlalchemy import func, and_, or_
from datetime import datetime, timedelta
import uuid
import io
import base64
import json

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import SessionLocal, AviationPart, ReportLog, ReportApprovalToken

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


# --- HELPERS ---

def get_db():
    """Get database session from Flask g context."""
    if 'db' not in g:
        g.db = SessionLocal()
    return g.db


def generate_report_id():
    """Generate unique aviation-compliant report ID."""
    date_str = datetime.utcnow().strftime('%Y%m%d')
    unique_suffix = uuid.uuid4().hex[:6].upper()
    return f"RPT-{date_str}-{unique_suffix}"


def apply_common_filters(query, args):
    """
    Apply cross-filtering parameters to query.
    
    Supported filters:
    - location: comma-separated list of locations
    - status: comma-separated list of tag colors (YELLOW,GREEN,WHITE,RED)
    - date_from / date_to: date range for registration_date
    - category: free-text search on part_name (future: dedicated category field)
    """
    # Location filter (multi-select)
    locations = args.get('location')
    if locations and locations.lower() not in ['all', 'todos', '']:
        loc_list = [loc.strip() for loc in locations.split(',') if loc.strip()]
        if loc_list:
            query = query.filter(func.lower(AviationPart.location).in_(
                [loc.lower() for loc in loc_list]
            ))
    
    # Status/Tag Color filter
    statuses = args.get('status')
    if statuses and statuses.lower() not in ['all', 'todos', '']:
        status_list = [s.strip().upper() for s in statuses.split(',') if s.strip()]
        if status_list:
            query = query.filter(AviationPart.tag_color.in_(status_list))
    
    # Date range filter
    date_from = args.get('date_from')
    date_to = args.get('date_to')
    if date_from and date_from.strip():
        query = query.filter(AviationPart.registration_date >= date_from)
    if date_to and date_to.strip():
        query = query.filter(AviationPart.registration_date <= date_to)
    
    # Category filter (text search on part_name)
    category = args.get('category')
    if category and category.lower() not in ['all', 'todos', '']:
        cat_list = [c.strip() for c in category.split(',') if c.strip()]
        if cat_list:
            conditions = [AviationPart.part_name.ilike(f'%{cat}%') for cat in cat_list]
            query = query.filter(or_(*conditions))
    
    return query


def part_to_report_dict(part):
    """Convert AviationPart model to dictionary for reports."""
    return {
        'id': part.id,
        'tagColor': part.tag_color,
        'partName': part.part_name,
        'brand': part.brand,
        'model': part.model,
        'pn': part.pn,
        'sn': part.sn,
        'ttTat': part.tt_tat,
        'tso': part.tso,
        'trem': part.trem,
        'tc': part.tc,
        'cso': part.cso,
        'crem': part.crem,
        'registrationDate': part.registration_date,
        'location': part.location,
        'organization': part.organization,
        'shelfLife': part.shelf_life,
        'removalReason': part.removal_reason,
        'technicalReport': part.technical_report,
        'removedFromAc': part.removed_from_ac,
        'position': part.position,
        'physicalStorageLocation': part.physical_storage_location,
        'rejectionReason': part.rejection_reason,
        'finalDisposition': part.final_disposition,
        'observations': part.observations,
        'technicianName': part.technician_name,
        'technicianLicense': part.technician_license,
        'inspectorName': part.inspector_name,
        'inspectorLicense': part.inspector_license,
        'signedByTechnician': part.signed_by_technician,
        'signedByInspector': part.signed_by_inspector,
    }


def build_report_response(report_type, data, filters_applied, user_name=None):
    """
    Build standardized report response with metadata.
    """
    # Calculate summary statistics
    status_counts = {'YELLOW': 0, 'GREEN': 0, 'WHITE': 0, 'RED': 0}
    for item in data:
        tag = item.get('tagColor')
        if tag in status_counts:
            status_counts[tag] += 1
    
    total = len(data)
    percentages = {}
    if total > 0:
        for status, count in status_counts.items():
            percentages[status] = round((count / total) * 100, 2)
    
    return {
        'reportId': generate_report_id(),
        'reportType': report_type,
        'generatedAt': datetime.utcnow().isoformat() + 'Z',
        'generatedBy': user_name or 'System',
        'filtersApplied': filters_applied,
        'data': data,
        'summary': {
            'total': total,
            'byStatus': status_counts,
            'percentages': percentages
        }
    }



def log_report_activity(activity_type, report_id, report_type, filters, details=None):
    """
    Log report related activity to database for audit trail.
    
    Args:
        activity_type: 'GENERATE', 'DOWNLOAD_PDF', 'DOWNLOAD_EXCEL', 'EMAIL'
        report_id: The ID of the report
        report_type: Category (TOTAL_INVENTORY, BY_STATUS, etc)
        filters: JSON serializable dict of filters applied
        details: Optional string with extra info (e.g. recipient email)
    """
    try:
        user_id = getattr(request, 'user_id', 'system')
        db = get_db()
        
        # Get user name for denormalization
        from .models import User
        user = db.query(User).filter(User.id == user_id).first()
        user_name = user.name if user else 'Unknown'
        
        # Create log entry
        new_log = ReportLog(
            id=str(uuid.uuid4()),
            report_id=report_id,
            report_type=report_type,
            user_id=user_id,
            user_name=user_name,
            acknowledge_timestamp=datetime.utcnow(), # Used as created_at for action log
            device_fingerprint=request.headers.get('User-Agent', 'unknown')[:255],
            recipient_email=details if activity_type == 'EMAIL' else None,
            item_count=str(activity_type), # Overloading field for action type if schema is rigid
            filters_applied=filters,
            status=activity_type, # Using status field for Action Type
            created_at=datetime.utcnow()
        )
        
        db.add(new_log)
        db.commit()
    except Exception as e:
        print(f"FAILED TO LOG AUDIT: {e}") # Fallback to stdout


# --- MIDDLEWARE FOR TOKEN ---
# Import token_required from main module dynamically to avoid circular imports
def get_current_user():
    """Get current user from request context (set by token_required)."""
    from .models import User
    db = get_db()
    user_id = getattr(request, 'user_id', None)
    if user_id:
        user = db.query(User).filter(User.id == user_id).first()
        return user.name if user else 'Unknown'
    return 'Unknown'


# We'll use a simple decorator that checks for token
def token_required_reports(f):
    """Token validation decorator for reports endpoints."""
    from functools import wraps
    from .models import UserSession
    
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({"message": "Token de seguridad no proporcionado"}), 401
        
        token = auth_header.split(' ')[1]
        db = get_db()
        session_record = db.query(UserSession).filter(UserSession.token == token).first()
        
        if not session_record or session_record.expiry < datetime.utcnow():
            if session_record:
                db.delete(session_record)
                db.commit()
            return jsonify({"message": "Sesión inválida o expirada"}), 401
        
        request.user_id = session_record.user_id
        return f(*args, **kwargs)
    return decorated


# --- REPORT ENDPOINTS ---

@reports_bp.route('/inventory', methods=['GET'])
@token_required_reports
def report_total_inventory():
    """
    Full Inventory Report (Global Fleet/Stock)
    
    Complete view of all assets including P/N, S/N, description,
    airworthiness status, and current location.
    
    Query Parameters:
    - location: comma-separated locations
    - status: comma-separated tag colors
    - date_from / date_to: date range
    - category: free-text category search
    """
    db = get_db()
    query = db.query(AviationPart)
    query = apply_common_filters(query, request.args)
    
    # Order by registration date descending (most recent first)
    query = query.order_by(AviationPart.registration_date.desc())
    
    parts = query.all()
    data = [part_to_report_dict(p) for p in parts]
    
    filters = {
        'location': request.args.get('location'),
        'status': request.args.get('status'),
        'dateFrom': request.args.get('date_from'),
        'dateTo': request.args.get('date_to'),
        'category': request.args.get('category')
    }
    # Remove None values
    filters = {k: v for k, v in filters.items() if v}
    
    response_data = build_report_response(
        'TOTAL_INVENTORY',
        data,
        filters,
        get_current_user()
    )
    
    # AUDIT LOG
    log_report_activity('GENERATE', response_data['reportId'], 'TOTAL_INVENTORY', filters)
    
    return jsonify(response_data)


@reports_bp.route('/by-status', methods=['GET'])
@token_required_reports
def report_by_status():
    """
    Report by Card Type (Status Compliance)
    
    Groups inventory by tag color: YELLOW (Serviceable), WHITE (Removed),
    RED (Rejected), GREEN (Repairable). Includes counts and availability percentages.
    """
    db = get_db()
    query = db.query(AviationPart)
    query = apply_common_filters(query, request.args)
    
    parts = query.all()
    
    # Group by status
    grouped_data = {
        'YELLOW': [],
        'GREEN': [],
        'WHITE': [],
        'RED': []
    }
    
    for part in parts:
        part_dict = part_to_report_dict(part)
        tag = part.tag_color
        if tag in grouped_data:
            grouped_data[tag].append(part_dict)
    
    filters = {
        'location': request.args.get('location'),
        'dateFrom': request.args.get('date_from'),
        'dateTo': request.args.get('date_to'),
        'category': request.args.get('category')
    }
    filters = {k: v for k, v in filters.items() if v}
    
    # Calculate totals per status
    total = len(parts)
    summary = {
        'total': total,
        'byStatus': {status: len(items) for status, items in grouped_data.items()},
        'percentages': {}
    }
    if total > 0:
        summary['percentages'] = {
            status: round((len(items) / total) * 100, 2)
            for status, items in grouped_data.items()
        }
    
    # AUDIT LOG
    log_report_activity('GENERATE', report_id, 'BY_STATUS', filters)
    
    return jsonify({
        'reportId': report_id,
        'reportType': 'BY_STATUS',
        'generatedAt': datetime.utcnow().isoformat() + 'Z',
        'generatedBy': get_current_user(),
        'filtersApplied': filters,
        'groupedData': grouped_data,
        'summary': summary
    })


@reports_bp.route('/by-location', methods=['GET'])
@token_required_reports
def report_by_location():
    """
    Report by Location (Logistics Layout)
    
    Stock breakdown based on physical infrastructure (Racks, Work Tables, Containers).
    Ideal for cyclic inventories.
    """
    db = get_db()
    query = db.query(AviationPart)
    query = apply_common_filters(query, request.args)
    
    parts = query.all()
    
    # Group by location
    location_groups = {}
    for part in parts:
        loc = part.location or 'Sin Ubicación'
        if loc not in location_groups:
            location_groups[loc] = {
                'items': [],
                'byStatus': {'YELLOW': 0, 'GREEN': 0, 'WHITE': 0, 'RED': 0}
            }
        location_groups[loc]['items'].append(part_to_report_dict(part))
        tag = part.tag_color
        if tag in location_groups[loc]['byStatus']:
            location_groups[loc]['byStatus'][tag] += 1
    
    # Calculate location summary
    location_summary = []
    for loc_name, loc_data in location_groups.items():
        location_summary.append({
            'location': loc_name,
            'totalItems': len(loc_data['items']),
            'byStatus': loc_data['byStatus']
        })
    
    # Sort by total items descending
    location_summary.sort(key=lambda x: x['totalItems'], reverse=True)
    
    filters = {
        'status': request.args.get('status'),
        'dateFrom': request.args.get('date_from'),
        'dateTo': request.args.get('date_to'),
        'category': request.args.get('category')
    }
    filters = {k: v for k, v in filters.items() if v}
    
    # AUDIT LOG
    log_report_activity('GENERATE', report_id, 'BY_LOCATION', filters)
    
    return jsonify({
        'reportId': report_id,
        'reportType': 'BY_LOCATION',
        'generatedAt': datetime.utcnow().isoformat() + 'Z',
        'generatedBy': get_current_user(),
        'filtersApplied': filters,
        'groupedData': location_groups,
        'locationSummary': location_summary,
        'summary': {
            'totalLocations': len(location_groups),
            'totalItems': len(parts)
        }
    })


@reports_bp.route('/by-pn/<string:part_number>', methods=['GET'])
@token_required_reports
def report_by_part_number(part_number):
    """
    Report by Part Number (P/N Performance)
    
    Deep analysis of a specific Part Number showing:
    - Total units in stock
    - All serial numbers
    - Status distribution
    - Movement history for each unit
    """
    db = get_db()
    
    # Find all parts with matching P/N (case-insensitive)
    query = db.query(AviationPart).filter(
        func.lower(AviationPart.pn) == part_number.lower()
    )
    
    # Apply other filters (except P/N-related)
    locations = request.args.get('location')
    if locations:
        loc_list = [loc.strip().lower() for loc in locations.split(',')]
        query = query.filter(func.lower(AviationPart.location).in_(loc_list))
    
    statuses = request.args.get('status')
    if statuses:
        status_list = [s.strip().upper() for s in statuses.split(',')]
        query = query.filter(AviationPart.tag_color.in_(status_list))
    
    parts = query.all()
    
    if not parts:
        return jsonify({
            'error': 'P/N not found in inventory',
            'message': f'Part Number "{part_number}" not found'
        }), 404
    
    # Build detailed response with serial numbers and history
    part_name = parts[0].part_name
    brand = parts[0].brand
    model = parts[0].model
    
    units = []
    status_counts = {'YELLOW': 0, 'GREEN': 0, 'WHITE': 0, 'RED': 0}
    
    for part in parts:
        unit_data = {
            'id': part.id,
            'sn': part.sn,
            'tagColor': part.tag_color,
            'location': part.location,
            'registrationDate': part.registration_date,
            'ttTat': part.tt_tat,
            'tso': part.tso,
            'trem': part.trem,
            'tc': part.tc,
            'cso': part.cso,
            'crem': part.crem,
            'shelfLife': part.shelf_life,
            'history': part.history or []
        }
        units.append(unit_data)
        
        if part.tag_color in status_counts:
            status_counts[part.tag_color] += 1
    
    total = len(parts)
    percentages = {}
    if total > 0:
        percentages = {
            status: round((count / total) * 100, 2)
            for status, count in status_counts.items()
        }
    
    filters = {
        'pn': part_number,
        'location': request.args.get('location'),
        'status': request.args.get('status')
    }
    filters = {k: v for k, v in filters.items() if v}
    
    # AUDIT LOG
    log_report_activity('GENERATE', report_id, 'BY_PART_NUMBER', filters)

    return jsonify({
        'reportId': report_id,
        'reportType': 'BY_PART_NUMBER',
        'generatedAt': datetime.utcnow().isoformat() + 'Z',
        'generatedBy': get_current_user(),
        'filtersApplied': filters,
        'partInfo': {
            'pn': part_number,
            'partName': part_name,
            'brand': brand,
            'model': model
        },
        'units': units,
        'summary': {
            'totalUnits': total,
            'byStatus': status_counts,
            'percentages': percentages
        }
    })


@reports_bp.route('/available-filters', methods=['GET'])
@token_required_reports
def get_available_filters():
    """
    Get available filter options based on current inventory data.
    
    Returns unique locations, tag colors in use, and date range.
    """
    db = get_db()
    
    # Get unique locations
    locations = db.query(AviationPart.location).distinct().filter(
        AviationPart.location != None,
        AviationPart.location != ''
    ).all()
    locations = sorted([loc[0] for loc in locations if loc[0]])
    
    # Get date range
    date_result = db.query(
        func.min(AviationPart.registration_date),
        func.max(AviationPart.registration_date)
    ).first()
    
    # Get tag colors in use
    tag_counts = db.query(
        AviationPart.tag_color,
        func.count(AviationPart.id)
    ).group_by(AviationPart.tag_color).all()
    
    return jsonify({
        'locations': locations,
        'tagColors': ['YELLOW', 'GREEN', 'WHITE', 'RED'],
        'tagColorCounts': dict(tag_counts),
        'dateRange': {
            'min': date_result[0] if date_result else None,
            'max': date_result[1] if date_result else None
        },
        'categories': [
            'ROTABLES',
            'CONSUMIBLES', 
            'MOTORES',
            'AVIONICS',
            'ESTRUCTURAL'
        ]
    })


# --- EMAIL DISPATCH ---

def generate_report_email_html(report_data):
    """
    Generate styled HTML email body following World Class Aviation industrial design.
    Uses table-based layout for email client compatibility.
    """
    report_type_labels = {
        'TOTAL_INVENTORY': 'Total Inventory / Inventario Total',
        'BY_STATUS': 'By Card Type / Por Tipo de Tarjeta',
        'BY_LOCATION': 'By Location / Por Ubicación',
        'BY_PART_NUMBER': 'By Part Number / Por Part Number'
    }
    
    report_type = report_data.get('reportType', 'UNKNOWN')
    summary = report_data.get('summary', {})
    filters = report_data.get('filtersApplied', {})
    
    # Build filters list
    filters_html = ''
    if filters:
        filters_list = ''.join([
            f'<tr><td style="padding: 4px 8px; font-size: 11px; color: #64748b;">{k}</td>'
            f'<td style="padding: 4px 8px; font-size: 11px; color: #1e293b; font-weight: bold;">{v}</td></tr>'
            for k, v in filters.items()
        ])
        filters_html = f'''
        <table cellpadding="0" cellspacing="0" style="width: 100%; margin-top: 12px;">
            <tr><td colspan="2" style="font-size: 10px; color: #94a3b8; text-transform: uppercase; padding-bottom: 4px;">Applied Filters / Filtros Aplicados</td></tr>
            {filters_list}
        </table>
        '''
    
    # Status breakdown with color indicators
    by_status = summary.get('byStatus', {})
    status_colors = {
        'YELLOW': '#eab308',
        'GREEN': '#10b981',
        'WHITE': '#64748b',
        'RED': '#f43f5e'
    }
    status_html = ''.join([
        f'''<td style="padding: 8px; text-align: center;">
            <div style="width: 12px; height: 12px; background-color: {status_colors.get(status, '#6366f1')}; 
                        border-radius: 2px; margin: 0 auto 4px auto;"></div>
            <div style="font-size: 18px; font-weight: 900; color: #0f172a;">{count}</div>
            <div style="font-size: 9px; color: #94a3b8; text-transform: uppercase;">{status}</div>
        </td>'''
        for status, count in by_status.items()
    ])
    
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #f8fafc;">
        <table cellpadding="0" cellspacing="0" style="width: 100%; max-width: 600px; margin: 0 auto; background-color: #ffffff;">
            
            <!-- HEADER: Black solid block -->
            <tr>
                <td style="background-color: #0f172a; padding: 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td>
                                <h1 style="margin: 0; color: #ffffff; font-size: 14px; font-weight: 900; 
                                           text-transform: uppercase; letter-spacing: 2px;">
                                    Aviation Technical Record
                                </h1>
                                <p style="margin: 4px 0 0 0; color: #94a3b8; font-size: 11px; text-transform: uppercase; letter-spacing: 1px;">
                                    Report Dispatch / Despacho de Reporte
                                </p>
                            </td>
                            <td style="text-align: right;">
                                <div style="background-color: #1e293b; padding: 8px 16px; border-radius: 4px;">
                                    <span style="color: #ffffff; font-size: 12px; font-weight: 900;">
                                        WORLD CLASS AVIATION
                                    </span>
                                </div>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- REPORT INFO -->
            <tr>
                <td style="padding: 24px 32px; border-bottom: 1px solid #e2e8f0;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td style="vertical-align: top; width: 50%;">
                                <p style="margin: 0; font-size: 10px; color: #94a3b8; text-transform: uppercase;">Report ID</p>
                                <p style="margin: 4px 0; font-size: 16px; font-weight: 900; color: #4f46e5; font-family: monospace;">
                                    {report_data.get('reportId', 'N/A')}
                                </p>
                            </td>
                            <td style="vertical-align: top; width: 50%; text-align: right;">
                                <p style="margin: 0; font-size: 10px; color: #94a3b8; text-transform: uppercase;">Generated</p>
                                <p style="margin: 4px 0; font-size: 12px; color: #1e293b; font-weight: bold;">
                                    {report_data.get('generatedAt', 'N/A')}
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- SUMMARY TABLE -->
            <tr>
                <td style="padding: 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%; border: 2px solid #0f172a;">
                        <!-- Table Header -->
                        <tr>
                            <td colspan="2" style="background-color: #0f172a; padding: 12px 16px;">
                                <span style="color: #ffffff; font-size: 11px; font-weight: 900; text-transform: uppercase; letter-spacing: 1px;">
                                    Report Summary / Resumen del Reporte
                                </span>
                            </td>
                        </tr>
                        <!-- Report Type -->
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc; border-bottom: 1px solid #e2e8f0; width: 40%;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Report Type / Tipo</span>
                            </td>
                            <td style="padding: 12px 16px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 13px; color: #0f172a; font-weight: 900;">
                                    {report_type_labels.get(report_type, report_type)}
                                </span>
                            </td>
                        </tr>
                        <!-- Item Count -->
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Item Count / Total Items</span>
                            </td>
                            <td style="padding: 12px 16px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 24px; color: #0f172a; font-weight: 900;">{summary.get('total', 0)}</span>
                                <span style="font-size: 11px; color: #94a3b8; margin-left: 8px;">units</span>
                            </td>
                        </tr>
                        <!-- Generated By -->
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Generated By / Por</span>
                            </td>
                            <td style="padding: 12px 16px;">
                                <span style="font-size: 13px; color: #0f172a; font-weight: bold;">
                                    {report_data.get('generatedBy', 'System')}
                                </span>
                            </td>
                        </tr>
                    </table>
                    
                    {filters_html}
                </td>
            </tr>
            
            <!-- STATUS BREAKDOWN -->
            <tr>
                <td style="padding: 0 32px 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%; background-color: #f8fafc; border: 1px solid #e2e8f0;">
                        <tr>
                            <td colspan="4" style="padding: 8px 12px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Status Breakdown / Desglose por Estado</span>
                            </td>
                        </tr>
                        <tr>
                            {status_html}
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- ATTACHMENT NOTE -->
            <tr>
                <td style="padding: 16px 32px; background-color: #fef3c7; border-top: 1px solid #fcd34d; border-bottom: 1px solid #fcd34d;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td style="width: 32px; vertical-align: top;">
                                <div style="width: 24px; height: 24px; background-color: #f59e0b; border-radius: 4px; text-align: center; line-height: 24px; color: white; font-weight: bold;">📎</div>
                            </td>
                            <td style="padding-left: 12px;">
                                <p style="margin: 0; font-size: 12px; color: #92400e; font-weight: bold;">
                                    Report Attachment / Adjunto del Reporte
                                </p>
                                <p style="margin: 4px 0 0 0; font-size: 11px; color: #a16207;">
                                    The complete report is attached to this email in the requested format.
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- FOOTER -->
            <tr>
                <td style="padding: 24px 32px; background-color: #0f172a;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td>
                                <p style="margin: 0; color: #64748b; font-size: 10px;">
                                    This is an automated message from the Aviation Inventory System.
                                </p>
                                <p style="margin: 4px 0 0 0; color: #475569; font-size: 10px;">
                                    World Class Aviation - Logistics Terminal Secure Layer
                                </p>
                            </td>
                            <td style="text-align: right;">
                                <p style="margin: 0; color: #94a3b8; font-size: 9px; font-family: monospace;">
                                    VER 6.0.1-EMAIL
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
        </table>
    </body>
    </html>
    '''
    
    return html


def generate_csv_content(report_data):
    """Generate CSV content from report data."""
    data = report_data.get('data') or []
    if report_data.get('groupedData'):
        data = []
        for items in report_data['groupedData'].values():
            data.extend(items)
    
    if not data:
        return "No data available"
    
    # CSV Headers
    headers = [
        'RECORD_ID', 'STATUS', 'P/N', 'S/N', 'PART_NAME', 'BRAND', 'MODEL',
        'LOCATION', 'TAT/T.T', 'TSO', 'T.REM', 'TC', 'CSO', 'C.REM',
        'REGISTRATION_DATE', 'SHELF_LIFE', 'OBSERVATIONS'
    ]
    
    lines = [','.join(headers)]
    
    for item in data:
        row = [
            str(item.get('id', '')),
            str(item.get('tagColor', '')),
            str(item.get('pn', '')),
            str(item.get('sn', '')),
            f'"{item.get("partName", "")}"',
            f'"{item.get("brand", "")}"',
            f'"{item.get("model", "")}"',
            f'"{item.get("location", "")}"',
            str(item.get('ttTat', '')),
            str(item.get('tso', '')),
            str(item.get('trem', '')),
            str(item.get('tc', '')),
            str(item.get('cso', '')),
            str(item.get('crem', '')),
            str(item.get('registrationDate', '')),
            str(item.get('shelfLife', '')),
            f'"{item.get("observations", "")}"'
        ]
        lines.append(','.join(row))
    
    return '\n'.join(lines)



def generate_pdf_report(report_data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    # Title
    report_type = report_data.get('reportType', 'Inventory Report')
    report_id = report_data.get('reportId', 'N/A')
    elements.append(Paragraph(f"World Class Aviation - {report_type}", title_style))
    elements.append(Paragraph(f"ID: {report_id} | Created: {report_data.get('generatedAt', '')}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Summary
    summary_text = f"Total Items: {report_data.get('summary', {}).get('totalItems', 0)} | User: {report_data.get('userName', 'System')}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Table Data
    data = []
    # Headers
    headers = ['Status', 'P/N', 'S/N', 'Part Name', 'Location', 'Condition', 'Qty']
    data.append(headers)
    
    items = report_data.get('items', [])
    if isinstance(items, dict): # Handle grouped data by flattening
        flat_items = []
        for group, group_items in items.items():
            flat_items.extend(group_items)
        items = flat_items

    for item in items:
        row = [
            item.get('status', 'PENDING'),
            item.get('partNumber', '')[:20],
            item.get('serialNumber', '')[:15],
            item.get('partName', '')[:25], # Truncate long names
            item.get('location', '')[:15],
            item.get('condition', '')[:10],
            str(item.get('quantity', 0))
        ]
        data.append(row)
        
    if len(data) < 2:
        elements.append(Paragraph("No items found in this report.", styles['Normal']))
    else:
        # Calculate column widths
        col_widths = [60, 100, 100, 200, 100, 60, 40]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Style
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.Color(0.2, 0.2, 0.2)), # Dark Gray
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ])
        table.setStyle(style)
        elements.append(table)
        
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(report_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Headers
    headers = ['Record ID', 'Status', 'P/N', 'S/N', 'Part Name', 'Location', 'Condition', 'Quantity', 'System Entry Date', 'Shelf Life', 'Observations']
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    items = report_data.get('items', [])
    if isinstance(items, dict): # Handle grouped data by flattening
        flat_items = []
        for group, group_items in items.items():
            flat_items.extend(group_items)
        items = flat_items

    for item in items:
        row = [
            item.get('id', ''),
            item.get('status', ''),
            item.get('partNumber', ''),
            item.get('serialNumber', ''),
            item.get('partName', ''),
            item.get('location', ''),
            item.get('condition', ''),
            item.get('quantity', 0),
            item.get('registrationDate', ''),
            item.get('shelfLife', ''),
            item.get('observations', '')
        ]
        ws.append(row)
        
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@reports_bp.route('/send-email', methods=['POST'])
@token_required_reports
def send_report_email():
    """
    Send report via email with PDF or CSV attachment.
    
    Request Body:
    {
        "recipients": ["email1@domain.com", "email2@domain.com"],
        "reportData": { ... },  // Report data from frontend
        "format": "PDF" | "CSV",
        "schedule": null | { "enabled": true, "frequency": "weekly" | "monthly" }
    }
    """
    from . import server_email
    import logging
    logger = logging.getLogger(__name__)
    
    data = request.json
    recipients = data.get('recipients', [])
    report_data = data.get('reportData', {})
    export_format = data.get('format', 'PDF')
    schedule = data.get('schedule')
    
    if not recipients:
        return jsonify({"message": "No recipients specified"}), 400
    
    if not report_data:
        return jsonify({"message": "No report data provided"}), 400
    
    # Generate email HTML body
    html_body = generate_report_email_html(report_data)
    
    # Subject line
    report_id = report_data.get('reportId', 'UNKNOWN')
    report_type = report_data.get('reportType', 'Report')
    subject = f"Aviation Technical Record - {report_type} - {report_id}"
    
    # Load email config
    cfg = server_email.load_config()
    
    # Generate attachment
    files = []
    try:
        if export_format == 'PDF':
            pdf_buffer = generate_pdf_report(report_data)
            files.append({
                'filename': f'{report_id}.pdf',
                'content': pdf_buffer.getvalue(),
                'mimetype': 'application/pdf'
            })
        elif export_format in ['EXCEL', 'XLSX']:
            xlsx_buffer = generate_excel_report(report_data)
            files.append({
                'filename': f'{report_id}.xlsx',
                'content': xlsx_buffer.getvalue(),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
        elif export_format == 'CSV':
            csv_content = generate_csv_content(report_data)
            files.append({
                'filename': f'{report_id}.csv',
                'content': csv_content.encode('utf-8'),
                'mimetype': 'text/csv'
            })
    except Exception as e:
        logger.error(f"Error generating attachment: {e}")

    # For now, we send to all recipients
    
    success_count = 0
    failed_recipients = []
    logs = []
    
    for recipient in recipients:
        try:
            success, msg = server_email.send_via_smtp(
                cfg, 
                recipient, 
                subject, 
                html_body,
                diagnostic_logs=logs,
                files=files
            )
            if success:
                success_count += 1
            else:
                failed_recipients.append(recipient)
                logger.error(f"Failed to send to {recipient}: {msg}")
        except Exception as e:
            failed_recipients.append(recipient)
            logger.error(f"Exception sending to {recipient}: {e}")
    
    # TODO: Handle schedule if provided
    if schedule and schedule.get('enabled'):
        # Store schedule in database for background processing
        logger.info(f"Report scheduling requested: {schedule.get('frequency')}")
        # This would integrate with a task scheduler like Celery or APScheduler
    
    if success_count == len(recipients):
        # AUDIT LOG
        log_report_activity('EMAIL', report_id, report_type, {}, details=f"Sent to {success_count} recipients (All)")

        return jsonify({
            "message": f"Reporte enviado exitosamente a {success_count} destinatario(s)",
            "success": True,
            "sent": success_count
        })
    elif success_count > 0:
        # AUDIT LOG
        log_report_activity('EMAIL', report_id, report_type, {}, details=f"Sent to {success_count} recipients (Partial)")

        return jsonify({
            "message": f"Reporte enviado a {success_count} de {len(recipients)} destinatarios",
            "success": True,
            "sent": success_count,
            "failed": failed_recipients
        })
    else:
        return jsonify({
            "message": "Error al enviar el reporte. Verifique la configuración SMTP.",
            "success": False,
            "failed": failed_recipients
        }), 500


@reports_bp.route('/download', methods=['POST'])
@token_required_reports
def download_report():
    data = request.json
    report_data = data.get('reportData', {})
    export_format = data.get('format', 'PDF').upper()
    
    if not report_data:
        return jsonify({"message": "No report data provided"}), 400
        
    report_id = report_data.get('reportId', 'report')
    filters = report_data.get('filtersApplied', {})
    
    # AUDIT LOG
    log_report_activity(f'DOWNLOAD_{export_format}', report_id, report_data.get('reportType', 'UNKNOWN'), filters)
    
    if export_format == 'PDF':
        buffer = generate_pdf_report(report_data)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{report_id}.pdf",
            mimetype='application/pdf'
        )
    elif export_format in ['EXCEL', 'XLSX']:
        buffer = generate_excel_report(report_data)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{report_id}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    elif export_format == 'CSV':
         # Reuse generate_csv_content but it returns string
         csv_content = generate_csv_content(report_data)
         return Response(
             csv_content,
             mimetype="text/csv",
             headers={"Content-disposition": f"attachment; filename={report_id}.csv"}
         )
         
    return jsonify({"message": f"Unsupported format: {export_format}"}), 400


# --- APPROVAL TOKEN SYSTEM ---

def generate_approval_token():
    """Generate a unique approval token."""
    return f"APR-{uuid.uuid4().hex}"


def create_device_fingerprint(request):
    """Create device fingerprint from request headers."""
    import hashlib
    user_agent = request.headers.get('User-Agent', 'unknown')
    ip = request.remote_addr or 'unknown'
    fingerprint_data = f"{user_agent}:{ip}"
    return hashlib.sha256(fingerprint_data.encode()).hexdigest()[:32]


def generate_approval_email_html(report_data, approval_token, base_url):
    """
    Generate styled HTML email with ACKNOWLEDGE & SIGN RECEIPT button.
    Follows World Class Aviation industrial design.
    """
    report_type_labels = {
        'TOTAL_INVENTORY': 'Total Inventory / Inventario Total',
        'BY_STATUS': 'By Card Type / Por Tipo de Tarjeta',
        'BY_LOCATION': 'By Location / Por Ubicación',
        'BY_PART_NUMBER': 'By Part Number / Por Part Number'
    }
    
    report_type = report_data.get('reportType', 'UNKNOWN')
    summary = report_data.get('summary', {})
    filters = report_data.get('filtersApplied', {})
    
    # Build filters list
    filters_html = ''
    if filters:
        filters_list = ''.join([
            f'<tr><td style="padding: 4px 8px; font-size: 11px; color: #64748b;">{k}</td>'
            f'<td style="padding: 4px 8px; font-size: 11px; color: #1e293b; font-weight: bold;">{v}</td></tr>'
            for k, v in filters.items()
        ])
        filters_html = f'''
        <table cellpadding="0" cellspacing="0" style="width: 100%; margin-top: 12px;">
            <tr><td colspan="2" style="font-size: 10px; color: #94a3b8; text-transform: uppercase; padding-bottom: 4px;">Applied Filters / Filtros Aplicados</td></tr>
            {filters_list}
        </table>
        '''
    
    # Status breakdown with color indicators
    by_status = summary.get('byStatus', {})
    status_colors = {
        'YELLOW': '#eab308',
        'GREEN': '#10b981',
        'WHITE': '#64748b',
        'RED': '#f43f5e'
    }
    status_html = ''.join([
        f'''<td style="padding: 8px; text-align: center;">
            <div style="width: 12px; height: 12px; background-color: {status_colors.get(status, '#6366f1')}; 
                        border-radius: 2px; margin: 0 auto 4px auto;"></div>
            <div style="font-size: 18px; font-weight: 900; color: #0f172a;">{count}</div>
            <div style="font-size: 9px; color: #94a3b8; text-transform: uppercase;">{status}</div>
        </td>'''
        for status, count in by_status.items()
    ])
    
    # Approval button URL
    approval_url = f"{base_url}/api/reports/acknowledge/{approval_token}"
    
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #f8fafc;">
        <table cellpadding="0" cellspacing="0" style="width: 100%; max-width: 600px; margin: 0 auto; background-color: #ffffff;">
            
            <!-- HEADER: Black solid block -->
            <tr>
                <td style="background-color: #0f172a; padding: 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td>
                                <h1 style="margin: 0; color: #ffffff; font-size: 14px; font-weight: 900; 
                                           text-transform: uppercase; letter-spacing: 2px;">
                                    Aviation Technical Record
                                </h1>
                                <p style="margin: 4px 0 0 0; color: #94a3b8; font-size: 11px; text-transform: uppercase; letter-spacing: 1px;">
                                    Report Dispatch - Approval Required
                                </p>
                            </td>
                            <td style="text-align: right;">
                                <div style="background-color: #1e293b; padding: 8px 16px; border-radius: 4px;">
                                    <span style="color: #ffffff; font-size: 12px; font-weight: 900;">
                                        WORLD CLASS AVIATION
                                    </span>
                                </div>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- REPORT INFO -->
            <tr>
                <td style="padding: 24px 32px; border-bottom: 1px solid #e2e8f0;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td style="vertical-align: top; width: 50%;">
                                <p style="margin: 0; font-size: 10px; color: #94a3b8; text-transform: uppercase;">Report ID</p>
                                <p style="margin: 4px 0; font-size: 16px; font-weight: 900; color: #4f46e5; font-family: monospace;">
                                    {report_data.get('reportId', 'N/A')}
                                </p>
                            </td>
                            <td style="vertical-align: top; width: 50%; text-align: right;">
                                <p style="margin: 0; font-size: 10px; color: #94a3b8; text-transform: uppercase;">Generated</p>
                                <p style="margin: 4px 0; font-size: 12px; color: #1e293b; font-weight: bold;">
                                    {report_data.get('generatedAt', 'N/A')}
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- SUMMARY TABLE -->
            <tr>
                <td style="padding: 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%; border: 2px solid #0f172a;">
                        <tr>
                            <td colspan="2" style="background-color: #0f172a; padding: 12px 16px;">
                                <span style="color: #ffffff; font-size: 11px; font-weight: 900; text-transform: uppercase; letter-spacing: 1px;">
                                    Report Summary / Resumen del Reporte
                                </span>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc; border-bottom: 1px solid #e2e8f0; width: 40%;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Report Type / Tipo</span>
                            </td>
                            <td style="padding: 12px 16px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 13px; color: #0f172a; font-weight: 900;">
                                    {report_type_labels.get(report_type, report_type)}
                                </span>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Item Count / Total Items</span>
                            </td>
                            <td style="padding: 12px 16px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 24px; color: #0f172a; font-weight: 900;">{summary.get('total', 0)}</span>
                                <span style="font-size: 11px; color: #94a3b8; margin-left: 8px;">units</span>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 12px 16px; background-color: #f8fafc;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Generated By / Por</span>
                            </td>
                            <td style="padding: 12px 16px;">
                                <span style="font-size: 13px; color: #0f172a; font-weight: bold;">
                                    {report_data.get('generatedBy', 'System')}
                                </span>
                            </td>
                        </tr>
                    </table>
                    
                    {filters_html}
                </td>
            </tr>
            
            <!-- STATUS BREAKDOWN -->
            <tr>
                <td style="padding: 0 32px 24px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%; background-color: #f8fafc; border: 1px solid #e2e8f0;">
                        <tr>
                            <td colspan="4" style="padding: 8px 12px; border-bottom: 1px solid #e2e8f0;">
                                <span style="font-size: 10px; color: #64748b; text-transform: uppercase;">Status Breakdown / Desglose por Estado</span>
                            </td>
                        </tr>
                        <tr>
                            {status_html}
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- ACKNOWLEDGE BUTTON -->
            <tr>
                <td style="padding: 16px 32px 32px 32px;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td align="center">
                                <a href="{approval_url}" 
                                   style="display: inline-block; padding: 20px 48px; 
                                          background-color: #eab308; color: #0f172a; 
                                          font-size: 14px; font-weight: 900; text-transform: uppercase;
                                          text-decoration: none; letter-spacing: 1px;
                                          border: 3px solid #10b981; border-radius: 8px;
                                          box-shadow: 0 4px 12px rgba(234, 179, 8, 0.4);">
                                    ✓ ACKNOWLEDGE &amp; SIGN RECEIPT
                                </a>
                            </td>
                        </tr>
                        <tr>
                            <td align="center" style="padding-top: 12px;">
                                <p style="margin: 0; font-size: 10px; color: #94a3b8;">
                                    Click button above to confirm receipt of this report.
                                    This action will be recorded in the Aircraft Maintenance Log.
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
            <!-- FOOTER -->
            <tr>
                <td style="padding: 24px 32px; background-color: #0f172a;">
                    <table cellpadding="0" cellspacing="0" style="width: 100%;">
                        <tr>
                            <td>
                                <p style="margin: 0; color: #64748b; font-size: 10px;">
                                    This is an automated message from the Aviation Inventory System.
                                </p>
                                <p style="margin: 4px 0 0 0; color: #475569; font-size: 10px;">
                                    World Class Aviation - Logistics Terminal Secure Layer
                                </p>
                            </td>
                            <td style="text-align: right;">
                                <p style="margin: 0; color: #94a3b8; font-size: 9px; font-family: monospace;">
                                    VER 7.0.0-APPROVAL
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            
        </table>
    </body>
    </html>
    '''
    
    return html


def generate_acknowledgment_page_html(report_data, acknowledged_by):
    """
    Generate HTML landing page for successful report acknowledgment.
    Industrial World Class Aviation styling.
    """
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Report Acknowledgment - World Class Aviation</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ 
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                padding: 20px;
            }}
            .container {{
                background: white;
                border-radius: 24px;
                max-width: 500px;
                width: 100%;
                overflow: hidden;
                box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.5);
            }}
            .header {{
                background: #0f172a;
                padding: 24px 32px;
                text-align: center;
            }}
            .header h1 {{
                color: white;
                font-size: 14px;
                font-weight: 900;
                text-transform: uppercase;
                letter-spacing: 2px;
                margin-bottom: 4px;
            }}
            .header p {{
                color: #94a3b8;
                font-size: 11px;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            .success-icon {{
                display: flex;
                justify-content: center;
                padding: 40px 32px 24px;
            }}
            .checkmark {{
                width: 100px;
                height: 100px;
                background: linear-gradient(135deg, #10b981 0%, #059669 100%);
                border-radius: 50%;
                display: flex;
                align-items: center;
                justify-content: center;
                box-shadow: 0 10px 40px rgba(16, 185, 129, 0.4);
                animation: pulse 2s infinite;
            }}
            @keyframes pulse {{
                0%, 100% {{ transform: scale(1); }}
                50% {{ transform: scale(1.05); }}
            }}
            .checkmark svg {{
                width: 50px;
                height: 50px;
                fill: white;
            }}
            .content {{
                padding: 0 32px 32px;
                text-align: center;
            }}
            .title {{
                font-size: 20px;
                font-weight: 900;
                color: #0f172a;
                text-transform: uppercase;
                letter-spacing: 1px;
                margin-bottom: 16px;
            }}
            .message {{
                font-size: 13px;
                color: #64748b;
                line-height: 1.6;
                margin-bottom: 24px;
            }}
            .details {{
                background: #f8fafc;
                border: 2px solid #e2e8f0;
                border-radius: 12px;
                padding: 16px;
                text-align: left;
            }}
            .detail-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
                border-bottom: 1px solid #e2e8f0;
            }}
            .detail-row:last-child {{
                border-bottom: none;
            }}
            .detail-label {{
                font-size: 10px;
                color: #94a3b8;
                text-transform: uppercase;
                font-weight: bold;
            }}
            .detail-value {{
                font-size: 12px;
                color: #0f172a;
                font-weight: bold;
                font-family: monospace;
            }}
            .footer {{
                background: #0f172a;
                padding: 16px 32px;
                text-align: center;
            }}
            .footer p {{
                color: #64748b;
                font-size: 10px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Report Acknowledgment Successful</h1>
                <p>World Class Aviation • Aircraft Maintenance Log</p>
            </div>
            
            <div class="success-icon">
                <div class="checkmark">
                    <svg viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                        <path d="M9 16.17L4.83 12l-1.42 1.41L9 19 21 7l-1.41-1.41L9 16.17z"/>
                    </svg>
                </div>
            </div>
            
            <div class="content">
                <h2 class="title">Acknowledgment Recorded</h2>
                <p class="message">
                    Your acknowledgment has been recorded in the Aircraft Maintenance Log.
                    This confirmation is permanent and cannot be modified.
                </p>
                
                <div class="details">
                    <div class="detail-row">
                        <span class="detail-label">Report ID</span>
                        <span class="detail-value">{report_data.get('reportId', 'N/A')}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Report Type</span>
                        <span class="detail-value">{report_data.get('reportType', 'N/A')}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Generated</span>
                        <span class="detail-value">{report_data.get('generatedAt', 'N/A')}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Acknowledged By</span>
                        <span class="detail-value">{acknowledged_by}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Timestamp</span>
                        <span class="detail-value">{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</span>
                    </div>
                </div>
            </div>
            
            <div class="footer">
                <p>This record is now read-only for compliance purposes.</p>
            </div>
        </div>
    </body>
    </html>
    '''


@reports_bp.route('/send-with-approval', methods=['POST'])
@token_required_reports
def send_report_with_approval():
    """
    Send report via email with approval token and ACKNOWLEDGE button.
    Creates ReportApprovalToken for tracking.
    
    Request Body:
    {
        "recipients": ["email1@domain.com"],
        "reportData": { ... },
        "format": "PDF" | "CSV"
    }
    """
    from . import server_email
    import logging
    import os
    logger = logging.getLogger(__name__)
    
    data = request.json
    recipients = data.get('recipients', [])
    report_data = data.get('reportData', {})
    
    if not recipients:
        return jsonify({"message": "No recipients specified"}), 400
    
    if not report_data:
        return jsonify({"message": "No report data provided"}), 400
    
    db = get_db()
    current_user_name = get_current_user()
    current_user_id = getattr(request, 'user_id', 'unknown')
    
    # Get base URL for approval links
    base_url = os.environ.get('BASE_URL', request.host_url.rstrip('/'))
    
    cfg = server_email.load_config()
    success_count = 0
    failed_recipients = []
    tokens_created = []
    
    for recipient in recipients:
        try:
            # Generate unique approval token
            token = generate_approval_token()
            
            # Store token in database
            approval_token = ReportApprovalToken(
                token=token,
                report_id=report_data.get('reportId', generate_report_id()),
                report_type=report_data.get('reportType', 'UNKNOWN'),
                report_data=report_data,
                recipient_email=recipient,
                sent_by_user_id=current_user_id,
                sent_by_user_name=current_user_name,
                created_at=datetime.utcnow(),
                expires_at=datetime.utcnow() + timedelta(days=7),
                acknowledged=False
            )
            db.add(approval_token)
            db.commit()
            tokens_created.append(token)
            
            # Generate email with approval button
            html_body = generate_approval_email_html(report_data, token, base_url)
            
            # Subject line
            report_id = report_data.get('reportId', 'UNKNOWN')
            report_type = report_data.get('reportType', 'Report')
            subject = f"[ACTION REQUIRED] Aviation Report - {report_type} - {report_id}"
            
            success, msg = server_email.send_via_smtp(
                cfg, 
                recipient, 
                subject, 
                html_body
            )
            
            if success:
                success_count += 1
            else:
                failed_recipients.append(recipient)
                logger.error(f"Failed to send to {recipient}: {msg}")
                
        except Exception as e:
            failed_recipients.append(recipient)
            logger.error(f"Exception sending to {recipient}: {e}")
            db.rollback()
    
    if success_count == len(recipients):
        return jsonify({
            "message": f"Reporte enviado con solicitud de aprobación a {success_count} destinatario(s)",
            "success": True,
            "sent": success_count,
            "tokens": tokens_created
        })
    elif success_count > 0:
        return jsonify({
            "message": f"Reporte enviado a {success_count} de {len(recipients)} destinatarios",
            "success": True,
            "sent": success_count,
            "failed": failed_recipients
        })
    else:
        return jsonify({
            "message": "Error al enviar el reporte.",
            "success": False,
            "failed": failed_recipients
        }), 500


@reports_bp.route('/acknowledge/<string:token>', methods=['GET'])
def acknowledge_report(token):
    """
    Process report acknowledgment via email link.
    Does not require full authentication if user has active session.
    Returns HTML landing page on success.
    """
    from flask import make_response
    import logging
    logger = logging.getLogger(__name__)
    
    db = get_db()
    
    # Find the approval token
    approval = db.query(ReportApprovalToken).filter(
        ReportApprovalToken.token == token
    ).first()
    
    if not approval:
        return make_response('''
            <html><body style="font-family: sans-serif; text-align: center; padding: 50px;">
                <h1 style="color: #dc2626;">Invalid Token</h1>
                <p>This approval link is invalid or has expired.</p>
            </body></html>
        ''', 404)
    
    # Check if already acknowledged
    if approval.acknowledged:
        return make_response(generate_acknowledgment_page_html(
            approval.report_data,
            f"{approval.acknowledged_by_user_name} (Already Acknowledged)"
        ), 200)
    
    # Check if expired
    if approval.expires_at < datetime.utcnow():
        return make_response('''
            <html><body style="font-family: sans-serif; text-align: center; padding: 50px;">
                <h1 style="color: #dc2626;">Token Expired</h1>
                <p>This approval link has expired. Please request a new report.</p>
            </body></html>
        ''', 410)
    
    # Get user from session if available
    auth_header = request.headers.get('Authorization')
    user_name = "Email Recipient"
    user_id = "email_recipient"
    
    if auth_header and auth_header.startswith('Bearer '):
        from .models import UserSession, User
        session_token = auth_header.split(' ')[1]
        session = db.query(UserSession).filter(UserSession.token == session_token).first()
        if session and session.expiry > datetime.utcnow():
            user = db.query(User).filter(User.id == session.user_id).first()
            if user:
                user_name = user.name
                user_id = user.id
    
    # Create device fingerprint
    fingerprint = create_device_fingerprint(request)
    
    # Update approval record
    approval.acknowledged = True
    approval.acknowledged_at = datetime.utcnow()
    approval.acknowledged_by_user_id = user_id
    approval.acknowledged_by_user_name = user_name
    approval.device_fingerprint = fingerprint
    
    # Create audit log entry (READ-ONLY after creation)
    log_entry = ReportLog(
        id=f"LOG-{uuid.uuid4().hex[:12].upper()}",
        report_id=approval.report_id,
        report_type=approval.report_type,
        user_id=user_id,
        user_name=user_name,
        acknowledge_timestamp=datetime.utcnow(),
        device_fingerprint=fingerprint,
        recipient_email=approval.recipient_email,
        item_count=str(approval.report_data.get('summary', {}).get('total', 0)),
        filters_applied=approval.report_data.get('filtersApplied', {}),
        status='ACKNOWLEDGED',
        created_at=datetime.utcnow()
    )
    db.add(log_entry)
    
    try:
        db.commit()
        logger.info(f"Report {approval.report_id} acknowledged by {user_name}")
    except Exception as e:
        db.rollback()
        logger.error(f"Error acknowledging report: {e}")
        return make_response('''
            <html><body style="font-family: sans-serif; text-align: center; padding: 50px;">
                <h1 style="color: #dc2626;">Error</h1>
                <p>An error occurred while processing your acknowledgment.</p>
            </body></html>
        ''', 500)
    
    # Return success landing page
    return make_response(
        generate_acknowledgment_page_html(approval.report_data, user_name),
        200
    )


@reports_bp.route('/logs', methods=['GET'])
@token_required_reports
def get_report_logs():
    """
    Get audit log entries. Read-only endpoint for compliance review.
    
    Query Parameters:
    - report_id: Filter by specific report
    - user_id: Filter by user
    - limit: Number of results (default 50)
    """
    db = get_db()
    
    query = db.query(ReportLog).order_by(ReportLog.acknowledge_timestamp.desc())
    
    report_id = request.args.get('report_id')
    if report_id:
        query = query.filter(ReportLog.report_id == report_id)
    
    user_id = request.args.get('user_id')
    if user_id:
        query = query.filter(ReportLog.user_id == user_id)
    
    limit = int(request.args.get('limit', 50))
    logs = query.limit(limit).all()
    
    return jsonify({
        'logs': [{
            'id': log.id,
            'reportId': log.report_id,
            'reportType': log.report_type,
            'userId': log.user_id,
            'userName': log.user_name,
            'acknowledgeTimestamp': log.acknowledge_timestamp.isoformat() if log.acknowledge_timestamp else None,
            'deviceFingerprint': log.device_fingerprint,
            'recipientEmail': log.recipient_email,
            'itemCount': log.item_count,
            'filtersApplied': log.filters_applied,
            'status': log.status,
            'createdAt': log.created_at.isoformat() if log.created_at else None
        } for log in logs],
        'count': len(logs)
    })


@reports_bp.route('/tokens', methods=['GET'])
@token_required_reports
def get_report_tokens():
    """
    Get all report approval tokens for the history dashboard.
    
    Query Parameters:
    - status: 'pending' | 'acknowledged' | 'all' (default: all)
    - type: report type filter
    - limit: Number of results (default 100)
    """
    db = get_db()
    
    query = db.query(ReportApprovalToken).order_by(ReportApprovalToken.created_at.desc())
    
    status = request.args.get('status', 'all').lower()
    if status == 'pending':
        query = query.filter(ReportApprovalToken.acknowledged == False)
    elif status == 'acknowledged':
        query = query.filter(ReportApprovalToken.acknowledged == True)
    
    report_type = request.args.get('type')
    if report_type:
        query = query.filter(ReportApprovalToken.report_type == report_type)
    
    limit = int(request.args.get('limit', 100))
    tokens = query.limit(limit).all()
    
    return jsonify({
        'tokens': [{
            'token': t.token,
            'reportId': t.report_id,
            'reportType': t.report_type,
            'recipientEmail': t.recipient_email,
            'sentByUserName': t.sent_by_user_name,
            'createdAt': t.created_at.isoformat() if t.created_at else None,
            'expiresAt': t.expires_at.isoformat() if t.expires_at else None,
            'acknowledged': t.acknowledged,
            'acknowledgedAt': t.acknowledged_at.isoformat() if t.acknowledged_at else None,
            'acknowledgedByUserName': t.acknowledged_by_user_name,
            'deviceFingerprint': t.device_fingerprint,
            'reportData': t.report_data
        } for t in tokens],
        'count': len(tokens)
    })



