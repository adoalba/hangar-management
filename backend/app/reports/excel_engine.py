
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .shared import format_report_items

logger = logging.getLogger(__name__)

def generate_excel_report(report_data):
    """
    Enterprise-Grade Excel Generator.
    Features: Frozen Headers, Filters, Brand Styling, Auto-sizing.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aviation Inventory Report"
        
        # Mandatory Formatter Usage
        report_type = report_data.get('reportType', 'TOTAL_INVENTORY')
        formatted = format_report_items(report_data.get('items', []), report_type)
        columns = formatted['columns']
        rows = formatted['rows']
        
        # Styles
        # Ink Efficient: White Background, Black Text
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Headers
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col in enumerate(columns, 1):
                val = row_data.get(col['key'], '—')
                ws.cell(row=row_idx, column=col_idx, value=val)
            
        # 1. Freeze Header
        ws.freeze_panes = "A2"
        
        # 2. Add Auto-Filters
        if columns:
            last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"
            
        # 3. Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 60)

        # 4. Zero-Row Safety
        if not rows:
            ws.cell(row=2, column=1, value="No records found for this criteria.")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns) if columns else 1)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"CRITICAL: Excel Generation Failure for {report_data.get('reportId')}: {e}")
        from ..utils.normalization import send_critical_alert
        send_critical_alert(f"Excel Generation Failure", f"Error building Excel workbook for report {report_data.get('reportId')}", report_id=report_data.get('reportId'), error=e)
        
        # Return fallback excel with error message
        try:
            eb = Workbook()
            es = eb.active
            es.append(["CRITICAL ERROR", "Failed to generate Excel report."])
            es.append(["Report ID", report_data.get('reportId', 'N/A')])
            es.append(["Details", "The system administrator has been notified."])
            
            err_buffer = io.BytesIO()
            eb.save(err_buffer)
            err_buffer.seek(0)
            return err_buffer
        except:
            return None

def generate_csv_content(report_data):
    """
    Enterprise-Grade CSV Generator.
    Features: UTF-8 with BOM for Excel compatibility, Stable labels.
    """
    try:
        report_type = report_data.get('reportType', 'TOTAL_INVENTORY')
        items = report_data.get('items', [])
        
        # MANDATORY FORMATTER
        formatted = format_report_items(items, report_type)
        columns = formatted['columns']
        rows = formatted['rows']

        # Headers
        headers = [col['label'] for col in columns]
        lines = [','.join(headers)]
        
        for row in rows:
            line_parts = []
            for col in columns:
                val = str(row.get(col['key'], '—'))
                if ',' in val or '"' in val or '\n' in val:
                    val = val.replace('"', '""')
                    val = f'"{val}"'
                line_parts.append(val)
            lines.append(','.join(line_parts))
        
        # Task 4: UTF-8 BOM
        content = '\n'.join(lines)
        return '\ufeff' + content # Add BOM
    except Exception as e:
        logger.error(f"CRITICAL: CSV Generation Failure: {e}")
        return "\ufeffERROR: Failed to generate CSV report content."
    except Exception as e:
        logger.error(f"CRITICAL: CSV Generation Failure for {report_data.get('reportId')}: {e}")
        from ..utils.normalization import send_critical_alert
        send_critical_alert(f"CSV Generation Failure", f"Error generating CSV for report {report_data.get('reportId')}", report_id=report_data.get('reportId'), error=e)
        return "ERROR: Failed to generate CSV report content. System administrator notified."
